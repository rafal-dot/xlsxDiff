# -*- coding: utf-8 -*-
"""
xlsxDiff.py: Excel .xlsx spreadsheet files comparison tool. It compares spreadsheets cell by cell and
produces output in Word changes tracking style on cell level

https://github.com/rafal-dot/xlsxDiff

Created on Sat Apr 9 19:26:10 2020

@author: Rafal Czeczotka <rafal dot czeczotka at gmail.com>

Copyright (c) 2020-2023 Rafal Czeczotka

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as published
by the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>."""

import argparse

from difflib import SequenceMatcher

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

import xlsxwriter


def log_print_message(is_log_enabled, message):
    """
    Print log message on screen

    :param is_log_enabled: defines, is message should be logged
    :param message: message text
    :return: current message length
    """
    if is_log_enabled:
        trim_len = log_print_message.log_msg_len - len(message) if len(message) < log_print_message.log_msg_len else 0
        print(message + " " * trim_len + "\r", end="")
        log_print_message.log_msg_len = len(message)


def get_format(row, column, modified_rows_set, modified_columns_set, f_do_nothing, f_highlight,
               do_update_modified_columns_rows_sets=False, column_row_added_deleted_format=None):
    """
    Updates sets of columns and rows, checks if cell should be highlighted and returns relevant format

    :param column: current column
    :param row: current row
    :param modified_rows_set: set of modified rows
    :param modified_columns_set: set of modified columns
    :param f_do_nothing: format to be returned, if cell should not be highlighted
    :param f_highlight: format to be returned, if cell should be highlighted
    :param do_update_modified_columns_rows_sets: information, if relevant column and row should be displayed as modified
    :param column_row_added_deleted_format: format for added/removed column(s)/row(s)
    :return: applicable format of cell
    """
    if do_update_modified_columns_rows_sets:
        modified_rows_set.add(row)
        modified_columns_set.add(column)
    if (column == 0 and row in modified_rows_set) or (row == 0 and column in modified_columns_set):
        return f_highlight
    if column_row_added_deleted_format is not None and args.no_highlight_added_removed:
        return column_row_added_deleted_format
    return f_do_nothing


def clone_tab(out_ws, tab_key, i_ws, operation_type):
    """
    Clone the entire contents of an input Excel tab to an output tab to present it as added or removed,
    and apply the formatting relevant to type

    :param out_ws: output worksheet
    :param tab_key: tab name
    :param i_ws: input worksheet
    :param operation_type: operation type: 'added' or 'removed'
    """
    tab_color, text_format = (None, None)
    if operation_type == "added":
        tab_color, text_format = ("blue", f_added)
    elif operation_type == "removed":
        tab_color, text_format = ("red", f_removed)
    for c in range(i_ws.max_column - 1, -1, -1):
        log_print_message(not args.quiet, f"New tab: {tab_key}  column: {c}")
        out_ws.set_column(c, c, i_ws.column_dimensions[get_column_letter(c + 1)].width)
        for r in range(i_ws.max_row - 1, -1, -1):
            log_print_message(not args.quiet and args.verbose, f"New tab: {tab_key}  column: {c}  row: {r}")
            v = i_ws.cell(r + 1, c + 1).value
            ret_val = str(v) if v is not None else ""
            out_ws.write(r, c, ret_val, text_format)
    out_ws.set_tab_color(tab_color)


def compare_cell(out_ws, o_r, o_c, in1_ws, i1_r, i1_c, in2_ws, i2_r, i2_c, modified_rows, modified_cols):
    """
    Compares two cells from 1st and 2nd worksheet and writes output in output worksheet

    :param out_ws: output worksheet
    :param o_r: row in output worksheet
    :param o_c: column in output worksheet
    :param in1_ws: 1st input worksheet
    :param i1_r: row in 1st input worksheet - if "-1"", then cell added
    :param i1_c: column in 1st input worksheet - if "-1"", then cell added
    :param in2_ws: 2nd input worksheet
    :param i2_r: row in 2nd input worksheet - if "-1"", then cell removed
    :param i2_c: column in 2nd input worksheet - if "-1"", then cell removed
    :param modified_rows: set of modified rows
    :param modified_cols: set of modified columns
    :return: returns True is cell is modified between 1st and 2nd spreadsheet and False elsewhere
    """
    v1 = "" if i1_r == -1 or i1_c == -1 else in1_ws.cell(i1_r + 1, i1_c + 1).value
    v2 = "" if i2_r == -1 or i2_c == -1 else in2_ws.cell(i2_r + 1, i2_c + 1).value
    v1 = "" if v1 is None else str(v1)
    v2 = "" if v2 is None else str(v2)

    # Check if cells are (i) empty, (ii) identical, (iii) new (cell or row/column) or (iv) removed (cell or row/column)
    if v1 == "" and v2 == "" \
            and i1_r >= 0 and i1_c >= 0 \
            and i2_r >= 0 and i2_c >= 0:
        # Both cells empty
        if args.noempty:
            return False
        out_ws.write(o_r, o_c, v2,
                     get_format(o_r, o_c, modified_rows, modified_cols, f_equal_cell, f_equal_cell_modified))
        return False
    elif v1 == v2 \
            and i1_r >= 0 and i1_c >= 0 \
            and i2_r >= 0 and i2_c >= 0:
        # Both cells equal, but not empty
        out_ws.write(o_r, o_c, v1,
                     get_format(o_r, o_c, modified_rows, modified_cols, f_equal_cell, f_equal_cell_modified))
        return False
    elif (i1_c == -1 and i2_r == -1) or (i1_r == -1 and i2_c == -1):
        # Added and removed cross, so "synthetic"/"technical" cell ;-)
        out_ws.write(o_r, o_c, "",
                     get_format(o_r, o_c, modified_rows, modified_cols, f_equal_cell, f_equal_cell_modified))
        return False
    elif i1_c == -1 or i1_r == -1:
        # First column/row is empty, so column/row added
        out_ws.write(
            o_r, o_c, v2,
            get_format(o_r, o_c, modified_rows, modified_cols, f_added, f_added_cell_modified,
                       do_update_modified_columns_rows_sets=args.highlight,
                       column_row_added_deleted_format=f_added_cell_highlight))
        return True
    elif i2_c == -1 or i2_r == -1:
        # Second column/row is empty, so column/row removed
        out_ws.write(
            o_r, o_c, v1,
            get_format(o_r, o_c, modified_rows, modified_cols, f_removed, f_removed_cell_modified,
                       do_update_modified_columns_rows_sets=args.highlight,
                       column_row_added_deleted_format=f_removed_cell_highlight))
        return True
    elif v1 == "":
        # First cell empty, so cell value added
        out_ws.write(o_r, o_c, v2,
                     get_format(o_r, o_c, modified_rows, modified_cols, f_added, f_added_cell_modified,
                                do_update_modified_columns_rows_sets=args.highlight))
        return True
    elif v2 == "":
        # Second cell empty, so cell value removed
        out_ws.write(o_r, o_c, v1,
                     get_format(o_r, o_c, modified_rows, modified_cols, f_removed, f_removed_cell_modified,
                                do_update_modified_columns_rows_sets=args.highlight))
        return True

    # None of multiple conditione above applicable, so compare content in detail
    rich_text = []
    is_output_rich_string = False
    for tag, i1, i2, j1, j2 in SequenceMatcher(None, v1, v2).get_opcodes():
        if tag == "equal":
            rich_text.append(v1[i1:i2])
        elif tag == "delete":
            rich_text.extend([f_removed, v1[i1:i2]])
            is_output_rich_string = True
        elif tag == "insert":
            rich_text.extend([f_added, v2[j1:j2]])
            is_output_rich_string = True
        elif tag == "replace":
            rich_text.extend([f_removed, v1[i1:i2], f_added, v2[j1:j2]])
            is_output_rich_string = True

    if is_output_rich_string:
        out_ws.write_rich_string(o_r, o_c, *rich_text,
                                 get_format(o_r, o_c, modified_rows, modified_cols, f_simple, f_modified_cell,
                                            do_update_modified_columns_rows_sets=args.highlight))
    else:
        # After multiple obvious checks before SequenceMatcher in the beginning, this should not happen, but if...
        out_ws.write(o_r, o_c, str(rich_text),
                     get_format(o_r, o_c, modified_rows, modified_cols, f_simple, f_modified_cell,
                                do_update_modified_columns_rows_sets=args.highlight))
        return False
    return True


def compare_2_lists_and_give_indexes_with_enumerator(list1, list2):
    """
    Compares two lists of merged cells (indexes) and returns list of 3 indexes:
    (1) index to output spreadsheet, (2) index to 1st input spreadsheet and (3) index to 2nd input spreadsheet

    :param list1: 1st list
    :param list2: 2nd list
    :return: list of ordered indexes: (i) descending index to output spreadsheet, (ii) descending index from 1st
    spreadsheet ("-1" is for elements added to 2nd spreadsheet) and (iii) descending index from 2nd
    spreadsheet ("-1" is for elements removed to 1st spreadsheet)
    """
    alternated_indexes = []
    for tag, i1, i2, j1, j2 in SequenceMatcher(None, list1, list2).get_opcodes():
        if tag == "equal":
            for i, j in zip(range(i1, i2), range(j1, j2)):
                alternated_indexes.append([i, j])
        elif tag == "delete":
            for i in range(i1, i2):
                alternated_indexes.append([i, -1])
        elif tag == "insert":
            for j in range(j1, j2):
                alternated_indexes.append([-1, j])
        elif tag == "replace":
            for i in range(i1, i2):
                alternated_indexes.append([i, -1])
            for j in range(j1, j2):
                alternated_indexes.append([-1, j])
    return_indexes = []
    for output_index, input_indexes in enumerate(alternated_indexes):  # reverse order
        return_indexes.insert(0, [output_index] + input_indexes)
    return return_indexes


def column_ranges(tab_key, i1_ws, i2_ws):
    """
    Returns an array of 3 indices, which are consecutive indices to columns in (i) output spreadsheet,
    (ii) 1st input speradsheet and (iii) 2nd input spreadsheet

    :param tab_key: tab name
    :param i1_ws: 1st spreadsheet
    :param i2_ws: 2nd spreadsheet
    :return: list of ordered indexes to columns: (i) descending index to output spreadsheet, (ii) descending
    index from 1st spreadsheet ("-1" is for elements added to 2nd spreadsheet) and (iii) descending index
    from 2nd spreadsheet ("-1" is for elements removed to 1st spreadsheet)
    """
    max1, max2 = (i1_ws.max_column, i2_ws.max_column)
    if tab_key in irows.keys():
        data1, data2 = ([], [])
        for col in range(max(max1, max2)):
            s1, s2 = ("", "")
            for row in irows[tab_key]:
                v1 = i1_ws.cell(row, col + 1).value
                v2 = i2_ws.cell(row, col + 1).value
                s1 = s1 + "|" if v1 is None else s1 + "|" + str(v1)  # "|" is to distingush ""+"data" vs "data"+""
                s2 = s2 + "|" if v2 is None else s2 + "|" + str(v2)
            if col < max1:
                data1.append(s1)
            if col < max2:
                data2.append(s2)
        return compare_2_lists_and_give_indexes_with_enumerator(data1, data2)
    else:
        return [[i, i, i] for i in range(max(max1, max2) - 1, -1, -1)]


def row_ranges(tab_key, i1_ws, i2_ws):
    """
    Returns an array of 3 indices, which are consecutive indices to rows in (i) output spreadsheet,
    (ii) 1st input speradsheet and (iii) 2nd input spreadsheet

    :param tab_key: tab name
    :param i1_ws: 1st spreadsheet
    :param i2_ws: 2nd spreadsheet
    :return: list of ordered indexes to rows: (i) descending index to output spreadsheet, (ii) descending
    index from 1st spreadsheet ("-1" is for elements added to 2nd spreadsheet) and (iii) descending index
    from 2nd spreadsheet ("-1" is for elements removed to 1st spreadsheet)
    """
    max1, max2 = (i1_ws.max_row, i2_ws.max_row)
    if tab_key in icolumns.keys():
        data1, data2 = ([], [])
        for row in range(max(max1, max2)):
            s1, s2 = ("", "")
            for col in icolumns[tab_key]:
                v1 = i1_ws.cell(row + 1, column_index_from_string(col)).value
                v2 = i2_ws.cell(row + 1, column_index_from_string(col)).value
                s1 = s1 + "|" if v1 is None else s1 + "|" + str(v1)  # "|" is to distingush ""+"data" vs "data"+""
                s2 = s2 + "|" if v2 is None else s2 + "|" + str(v2)
            if row < max1:
                data1.append(s1)
            if row < max2:
                data2.append(s2)
        return compare_2_lists_and_give_indexes_with_enumerator(data1, data2)
    else:
        return [[i, i, i] for i in range(max(max1, max2) - 1, -1, -1)]


def compare_tab(out_ws, tab_key, i1_ws, i2_ws):
    """
    Compare the contents of two Excel tabs (from 1st and 2nd input spreadsheets) and generate the contents
    of the output tab

    :param out_ws: output worksheet
    :param tab_key: tab name
    :param i1_ws: 1st input worksheet
    :param i2_ws: 2nd input worksheet
    """
    is_data_in_tab_modified = False

    modified_rows = set()
    modified_cols = set()
    row_ranges_list = row_ranges(tab_key, i1_ws, i2_ws)
    column_ranges_list = column_ranges(tab_key, i1_ws, i2_ws)
    for co, c1, c2 in column_ranges_list:
        log_print_message(not args.quiet, f"Comparing tab: \"{tab_key}\"  column: {co}")
        c1_width = i1_ws.column_dimensions[get_column_letter(c1 + 1)].width if c1 >= 0 else 2
        c2_width = i2_ws.column_dimensions[get_column_letter(c2 + 1)].width if c2 >= 0 else 2
        out_ws.set_column(co, co, max(c1_width, c2_width, 2))  # 2 for very narrow columns, to make them visible ;-)

        for ro, r1, r2 in row_ranges_list:
            log_print_message(not args.quiet and args.verbose,
                              f"Comparing tab: \"{tab_key}\"  column:  {co}  row:{ro:5}")
            compare_cell_retval = compare_cell(
                out_ws, ro, co,
                i1_ws, r1, c1, i2_ws, r2, c2,
                modified_rows, modified_cols)
            is_data_in_tab_modified = compare_cell_retval or is_data_in_tab_modified

    if not is_data_in_tab_modified:
        out_ws.set_tab_color("gray")
    elif args.autofilter and args.highlight:
        # column_ranges_list[0][0] is first element with 3 integer indexes (highest) and [0] item is the largest
        # index in output worksheet
        out_ws.autofilter(0, 0, 0, max(i1_ws.max_column, i2_ws.max_column, column_ranges_list[0][0] + 1) - 1)


parser = argparse.ArgumentParser(
    description="Compares two .xlsx Excel spreadsheets, cell content with cell content. Creates an output "
                "spreadsheet with all the detailed cell content differences, marking the changes in each cell "
                "in a format similar to the track changes feature in a Word word processor. More information "
                "can be found in the attached PDF documentation file. The latest version of xlsxDiff can always "
                "be found here: https://github.com/rafal-dot/xlsxDiff",
    epilog="Copyright (C) 2020-2023 Rafal Czeczotka. This program comes with ABSOLUTELY NO WARRANTY. "
           "This is free software, and you are welcome to redistribute it under GNU Affero GPL conditions "
           "(see https://www.gnu.org/licenses/).")
parser.add_argument("input1", help="first input spreadsheet file to compare", type=str)
parser.add_argument("input2", help="second input spreadsheet file to compare", type=str)
parser.add_argument("output", help="output spreadsheet file with highlighted differences", type=str)
parser.add_argument("-f", "--formula", help="compare formula text in cells instead of data values (default: disabled)",
                    action="store_true")
parser.add_argument(
    "-x", "--highlight",
    help="in each row, if there are any cell changed, mark the first cell with a green background (row 1:1). "
         "In each column, if there are any cell changed, mark the first cell with a green background (column A:A). "
         "This option is useful for large spreadsheets, to facilitate changes identification. To filter out chaged "
         "data, autofilter with color filter might be applied later in spreadsheet software (default: disabled)",
    action="store_true")
parser.add_argument(
    "-c", "--icolumn",
    help="provide information on index column(s) (example: \"-c Sheet1!B,C\")). This allows to identify entire "
         "added/deleted columns, making it much easier to identify changes. See the documentation for more details",
    action="append")
parser.add_argument("-r", "--irow",
                    help="provide information on index row(s) (example: \"-r Sheet2!1,2\"). "
                         "This allows to identify entire added/deleted rows, making it much easier to "
                         "identify changes. See the documentation for more details",
                    action="append")
parser.add_argument("-X", "--no_highlight_added_removed",
                    help="do not highlight added/removed columns/rows (see \"-r\" and \"-c\" options; "
                         "default: disabled)",
                    action="store_false")
parser.add_argument("-a", "--autofilter", help="add Excel's autofilter in changed tabs (default: disabled)",
                    action="store_true")
parser.add_argument("-e", "--noempty", help="ignore empty cells (default: disabled)", action="store_true")
parser.add_argument("-v", "--verbose", help="verbose output. As it takes time to process large spreadsheets, "
                                            "this option facilitates progress tracking", action="store_true")
parser.add_argument("-q", "--quiet", help="no output messages", action="store_true")
parser.add_argument("--version", action="version", version='%(prog)s 2.0.0 (2023-02-27)')
args = parser.parse_args()

icolumns = {}
if args.icolumn is not None:
    for icolumn in args.icolumn:
        ic_key, ic_columns = icolumn.split("!")
        icolumns[ic_key] = ic_columns.split(",")
irows = {}
if args.irow is not None:
    for irow in args.irow:
        ic_key, ic_row = irow.split("!")
        irows[ic_key] = [int(i) for i in ic_row.split(",")]

log_print_message.log_msg_len = 0
log_print_message(not args.quiet, f"Loading spreadsheet: {args.input1}")
i1_wb = load_workbook(filename=args.input1, data_only=not args.formula)
log_print_message(not args.quiet, f"Loading spreadsheet: {args.input2}")
i2_wb = load_workbook(filename=args.input2, data_only=not args.formula)

tab_names_dict = {}
for i2_tab in i2_wb.sheetnames:  # First, tabs from latest spreadsheet ;-)
    tab_names_dict[i2_tab] = {2, }
for i1_tab in i1_wb.sheetnames:
    if i1_tab in tab_names_dict.keys():
        tab_names_dict[i1_tab].add(1)
    else:
        tab_names_dict[i1_tab] = {1, }

if not args.quiet:
    for tab_name in set(icolumns.keys()) | set(irows.keys()):
        if tab_name not in tab_names_dict.keys():
            print(f"WARNING: There is no tab name \"{tab_name}\"! Typo?")

o_wb = xlsxwriter.Workbook(args.output)

# If your Python interpreter raises issue here, just replace text:
# "**f_common"
# with following text:
# ""text_wrap": True, "border": 1, "align": "top""
# Merging dictionaries is not 100% standardised between various versions of Python yet
f_common = {"text_wrap": True, "border": 1, "align": "top"}
f_simple = o_wb.add_format({**f_common})
f_added = o_wb.add_format({**f_common, "bold": True, "underline": True, "color": "blue"})
f_added_cell_highlight = o_wb.add_format(
    {**f_common, "bold": True, "underline": True, "color": "blue", "bg_color": "#e8e8ff"})  # light blue
f_added_cell_modified = o_wb.add_format(
    {**f_common, "bold": True, "underline": True, "color": "blue", "bg_color": "#a9d171"})  # light green
f_removed = o_wb.add_format({**f_common, "bold": True, "font_strikeout": True, "color": "red"})
f_removed_cell_highlight = o_wb.add_format(
    {**f_common, "bold": True, "font_strikeout": True, "color": "red", "bg_color": "#ffe8e8"})  # light red
f_removed_cell_modified = o_wb.add_format(
    {**f_common, "bold": True, "font_strikeout": True, "color": "red", "bg_color": "#a9d171"})  # light green
f_modified_cell = o_wb.add_format({**f_common, "bg_color": "#a9d171"})  # light green
f_equal_cell = o_wb.add_format({**f_common, "bg_color": "#c0c0c0"})
f_equal_cell_modified = o_wb.add_format({**f_common, "bg_color": "#a9d171"})

for current_tab_key in tab_names_dict.keys():
    o_ws = o_wb.add_worksheet(current_tab_key)
    if tab_names_dict[current_tab_key] == {1, 2}:
        # Tab name exists in both spreadsheets, individual cells are compared
        compare_tab(o_ws, current_tab_key, i1_wb[current_tab_key], i2_wb[current_tab_key])

    elif tab_names_dict[current_tab_key] == {2, }:
        # Tab name exists only in 2nd spreadsheet, i.e. the tab is new, so no comaprison needed, just copy data
        # from 2nd input spreadsheet to output and use "added" formatting
        clone_tab(o_ws, current_tab_key, i2_wb[current_tab_key], "added")

    elif tab_names_dict[current_tab_key] == {1, }:
        # Tab name exists only in 1st spreadsheet i.e. the tab is removed, so no comaprison needed, just copy data
        # from 1st input spreadsheet to output and use "removed" formatting
        clone_tab(o_ws, current_tab_key, i1_wb[current_tab_key], "removed")

log_print_message(not args.quiet, f"Saving spreadsheet: {args.output}")
o_wb.close()
log_print_message(not args.quiet, " ")
